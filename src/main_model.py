import gurobipy as gp
from gurobipy import GRB
from gurobipy import *
from inputs_data import *
from preprocesamiento import *
from GrafoFinal import GrafoColores
import xlwt
import openpyxl
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import pandas as pd
# -----------------------------------------------------------------
# ---------------------       INPUT DATA   ------------------------
# -----------------------------------------------------------------

user_variables = {

    'MinLen' : 1,
    'MaxLen' : 150,
    'S_min' : 4,
    'S_max' : 14,
    'preprocesamiento': True, #True,#False,#False,# or True
    'pathData' : "C:/Users/Fernando/Desktop/TFG/pythonMathModels/Data/",
    'pathSolutions' : "C:/Users/Fernando/Desktop/TFG/pythonMathModels/Soluciones/",
    'instance' : "Mandl",#"Mandl",
    'stop_file_name' :"Stops_Mandl.txt", #""Stops_old.txt", #"Stops_Mandl.txt",#
    'stopPED_file_name':"Stops_PED_Mandl.txt",#"Stops_PED_old.txt", #Stops_PED_Mandl.txt",
    'arcBus_file_name' :"Arcs_bus_Mandl_reescalado.txt", #Arcs_bus_Small.txt",#"Arcs_bus_old.txt", #"Arcs_bus_Mandl.txt",#"Arcs_bus_Mandl_reescalado.txt", #"Arcs_bus_Mandl.txt",
    'arcPED_file_name': "Arcs_PED_Mandl_reescalado.txt", #Arcs_PED_Small.txt",#"Arcs_PED_old.txt", #"Arcs_PED_Mandl.txt",# "Arcs_PED_Mandl_reescalado.txt",
    'OD_name' : "OD_Mandl.txt", #"OD_old.txt",#"OD_Mandl.txt",#
    'coordenadas_name' : "Coordenadas.txt",
  #  'headway_list' : [15, 30], #headways in minutes [15, 30],[3, 5, 10, 15, 30]----->No se está usando!!
   #'frequecies' : [2, 4],
    #'frequencies' : [2, 4],#[2, 4],[2, 4, 6, 12, 20]
    'v_bus' : 40,
    'v_PED' : 3,
    'M' : 500, #upperbound on the total length for line---antes 1500
    'cap_1' : 110, #capacidad bus tipo 1
    'cap_2' : 81,
    'numLineasList' : [3,5,8], #3,#--3
    'frequencies_list_Total': [[12,20],[2,4,6],[2, 4, 6, 12, 20]],  # frecuencias bajas, medias y todas ,[2,4,6],[2,4,6,12,20]
    'fleet_max_List': [6,10,15], # [50, 70, 100],  # 50, #flota máxima por hora--20   10, 15
    'fleet_1_max_List': [3,5,10], #[25, 35, 50], #flota máxima tipo 1por hora---5    ,5,10
    'fleet_2_max_List': [3,5,10], #[25, 35, 50] #flota máxima tipo 2 por hora---10    ,5,10

}
# -----------------     from inputs_data.py    --------------------
# -----------------------------------------------------------------


pathSoluciones=user_variables['pathSolutions']
instance=user_variables['instance']
pathData=user_variables['pathData']
stop_file_name=user_variables['stop_file_name']
stopPED_file_name=user_variables['stopPED_file_name']
arcBus_file_name=user_variables['arcBus_file_name']
arcPED_file_name=user_variables['arcPED_file_name']
OD_name=user_variables['OD_name']

Lista_nodos=Lectura_StopsFile(pathData,stop_file_name)#lista con los nodos..[0,1,2,...]
Lista_nodos_PED=Lectura_StopsFile(pathData,stopPED_file_name)#lista con los nodos..[0,1,2,...]

Lista_arcos_bus=Lectura_ArcsBusFile(pathData,arcBus_file_name,0)#lista con los arcos [(0,1), (0,2)...]

Lista_arcs_PED=Lectura_ArcsPEDFile(pathData,arcPED_file_name,0)#lista con las aristas PED [(0,1), (0,2)...]

Lista_OD=Lectura_ODFile(pathData,OD_name,0) #lista con el Id del par, Origen, Destino y Pasajeros

dicc_OD_demanda=Lectura_ODFile(pathData,OD_name,1) #Diccionario tipo; OD:demanda

dicc_arco_distancia=Lectura_ArcsBusFile(pathData,arcBus_file_name,3)# diccionario tipo arco: distancia
dicc_arco_distancia_PED=Lectura_ArcsPEDFile(pathData,arcPED_file_name,3)# diccionario tipo arco: distancia
# print("dicc_arco_distancia",dicc_arco_distancia)
# print("dicc_OD_demanda",dicc_OD_demanda)
# print("Lista_arcos_bus",Lista_arcos_bus)
# print("Lista_OD", Lista_OD)
# print("Lista_arcs_PED", Lista_arcs_PED)


v_bus=user_variables['v_bus']
v_PED=user_variables['v_PED']
S_min=user_variables['S_min']
S_max=user_variables['S_max']
M=user_variables['M']
cap_1=user_variables['cap_1']
cap_2=user_variables['cap_2']


#Datos sobre los que se van a crear los escenarios:

numLineasList=user_variables['numLineasList']
#fleet_max=user_variables['fleet_max']
#fleet_1_max=user_variables['fleet_1_max']
#fleet_2_max=user_variables['fleet_2_max']
frequencies_list_Total = user_variables['frequencies_list_Total']
fleet_max_List= user_variables['fleet_max_List']
fleet_1_max_List=user_variables['fleet_1_max_List']
fleet_2_max_List=user_variables['fleet_2_max_List']
# .....OJO.....DENTRO DE BUCLE HAY QUE CREAR UN CONTADOR DEL ESCENARIO
contadorEscenario=0
iii=0
for numLineas in numLineasList:
    NumLines = numLineas
    for frequencies_list in frequencies_list_Total:
        #print("fleet_1_max_List", fleet_1_max_List)
        iii = 0
        for fleet_max in fleet_max_List:
            freq_max = frequencies_list[len(frequencies_list) - 1]
            freq_min = frequencies_list[0]
            print("+++++++++++++++++++numLineas", numLineas)
            print("+++++++++++++++++++frequencies_list", frequencies_list)
            print("+++++++++++++++++++fleet_max", fleet_max)

            #print("iii", iii)
            fleet_1_max=fleet_1_max_List[iii]
            #print("fleet_1_max", fleet_1_max)
            fleet_2_max = fleet_2_max_List[iii]
            #print("iii", iii)
            iii=iii+1
            print("+++++++++++++++++++fleet_1_max", fleet_1_max)
            print("+++++++++++++++++++fleet_2_max", fleet_2_max)
            print("++++++++++++++++++++++++++++freq_min", freq_min)
            contadorEscenario=contadorEscenario+1
            # -----------------     Creation of dictionaries     --------------------
            # -----------------------------------------------------------------
            preprocessing=user_variables['preprocesamiento']

            # lista de arcos con las distancias; lista con arco y distancias
            Lista_arcos_distancias_bus = Lectura_ArcsBusFile(pathData, arcBus_file_name, 2)

            if preprocessing:
                # -----------------     Creation of dictionaries from PREPROCESSING    --------------------

               # print("Lista_arcos_distancias_bus; ",Lista_arcos_distancias_bus)
                # --diccionario tipo; parOD: lista cuyos eltos son listas con los nodos de cada path:
                dicc_parOD_listaPaths = CreaDiccionario_KshortestPaths(Lista_nodos, Lista_arcos_distancias_bus, Lista_OD)

                # --diccionario tipo; parOD: lista con todos los nodos implicados en todos los paths:
                dicc_parOD_listaNodos = CreaDiccionario_listaNodos_KshortestPaths(dicc_parOD_listaPaths)

                # --diccionario tipo; parOD: lista con todos los arcos implicados en todos los paths:
                dicc_parOD_listaArcos = CreaDiccionario_listaArcos_KshortestPaths(dicc_parOD_listaPaths)

            else: # -----------------     Creation of dictionaries without PREPROCESSING    --------------------
                # --diccionario tipo; parOD: lista con todos los nodos del grafo bus:
                dicc_parOD_listaNodos=CreaDiccionario_listaNodos(Lista_nodos,Lista_OD)
                print("444444444444444444444444")
                # --diccionario tipo; parOD: lista con todos los arcos del grafo bus:
                dicc_parOD_listaArcos=CreaDiccionario_listaArcos(Lista_arcos_bus, Lista_OD)



            # --diccionario tipo; parOD: lista con todos las aristas PED:
            dicc_parOD_listaAristasPED=CreaDiccionario_listaArcos(Lista_arcs_PED, Lista_OD)

            #print("dicc_parOD_listaArcos", dicc_parOD_listaArcos)

            # -----------------------------------------------------------------
            # ---------------------       VARIABLES    ------------------------
            # -----------------------------------------------------------------
            mod = gp.Model("Freq_design")

            # ---------------------       declaration of variables    ------------------------
            #-- desicion variables:
            hw = {}#headway of each line
            fr = {}#frequency of each line
            z = {}#required fleet
            s = {}#stops of each line
            u = {}#arcs of each line
            r = {}#flow
            t = {}#pedestrian arcs
            v = {}#transfer between lines
            w = {}#change from pedestrian to bus
            delta_1={} # selección del tipo 1 de bus
            delta_2={} # selección del tipo 2 de bus
            z_1={} #nb buses tipo 1
            z_2={}#nb buses tipo 2
            h={}# variable representando si la linea se activa a no@@@@@@@@ 6/6/2023
            #--  auxiliary variables:

            fr_k = {}#para definir headway como una c.l. de binarias
            #eps_k = {}#para linealizar el producto hw y fr_k
            psi={}#variable representando el producto de u y fr
            eta={}#variable representando el producto de v y h
            sigma={}#variable representando el producto de r y h
            rho={}#variable representado el producto w and h

            # ---------------------     creation of variables    ------------------------
            #-------- DESIGN:


            Lines = list(range(1, NumLines + 1))

            # ------6/6/2023  indicador de activacion de lineas
            for l in Lines:
                h[l] = mod.addVar(lb=0,ub=1.0, vtype=GRB.BINARY, name="h_%s" % (l))
            mod.update()


            # s_i^l = 1 si la parada i se activa para la linea l
            for i in Lista_nodos:
                for l in Lines:
                    s[i, l] = mod.addVar(lb=0, ub=1.0, vtype=GRB.BINARY, name="s_%s_%s" % (i,l))
            mod.update()

            # u_ij^l =1, si el arco (i,j) se activa en l
            for (i, j) in Lista_arcos_bus:
                 for l in Lines:
                     u[i, j, l] = mod.addVar(lb=0, ub=1.0, vtype=GRB.BINARY, name="u_%s_%s_%s" % (i, j, l))
            mod.update()

            #-------- FLOW:
            # r_ij^Ol =1, si el flujo del par O pasa por (i,j) de la linea l
            for (origen, destino) in dicc_parOD_listaArcos: #dicc tipo; parOD: listaArcos
                for (i,j) in dicc_parOD_listaArcos[(origen,destino)]:
                    for l in Lines:
                        r[(i, j), (origen, destino), l] = mod.addVar(lb=0, ub=1.0, vtype=GRB.BINARY, name="r_(%s,%s)_(%s,%s)_%s" % (i, j, origen, destino, l))

            mod.update()

            # t_ij^O =1, si el flujo del par O pasa por arco pedestre (i,j)
            for (origen, destino) in dicc_parOD_listaAristasPED: #dicc tipo; parOD: listaArcos
                for (i,j) in dicc_parOD_listaAristasPED[(origen,destino)]:
                   t[(i, j), (origen, destino)] = mod.addVar(lb=0, ub=1.0, vtype=GRB.BINARY, name="t_(%s,%s)_(%s,%s)" % (i, j, origen, destino))

            mod.update()
            # v_i^Oll' =1, si los pasajeros del par O transfieren de la linea l a la l' en parada i
            for (origen, destino) in dicc_parOD_listaNodos: #dicc tipo; parOD: listaNodos
                for i in dicc_parOD_listaNodos[(origen,destino)]: #cogemos un nodo de la lista
                    if i != origen and i != destino: #seleccionamos los que no son ni origen ni destino
                        for l1 in Lines:
                            for l2 in Lines:
                                if l1 != l2:
                                    v[i, (origen, destino), l1, l2] = mod.addVar(lb=0, ub=1.0, vtype=GRB.BINARY, name="v_%s_(%s,%s)_%s_%s" % (i, origen, destino, l1, l2))

            mod.update()

            # w_i^Ol =1, si los pasajeros del par O cambian del modo pedestre al modo bus en el nodo i de la linea l
            for (origen, destino) in dicc_parOD_listaNodos: #dicc tipo; parOD: listaNodos
                #print("origen, destino: ", origen, destino)
                for i in dicc_parOD_listaNodos[(origen,destino)]: #cogemos un nodo de la lista de bus
                    #print("dicc_parOD_listaNodos[(origen,destino)]: ",dicc_parOD_listaNodos[(origen,destino)])
                   # print(Lista_nodos_PED)
                    if Lista_nodos_PED.__contains__(i): #siempre que i esté en los pedestres
                        #print(i)
                        for l in Lines:
                            w[i, (origen, destino), l] = mod.addVar(lb=0, ub=1.0, vtype=GRB.BINARY, name="w_%s_(%s,%s)_%s" % (i, origen, destino, l))

            mod.update()

            #-------- FREQUENCIES AND HEADWAYS AND TYPE OF BUS:

            # h^l headway line l
            for l in Lines:
                hw[l] = mod.addVar(lb=0, vtype=GRB.INTEGER, name="hw_%s" % (l))
            mod.update()
            # fr^l frequency line l
            for l in Lines:
                fr[l] = mod.addVar(lb=0, vtype=GRB.INTEGER, name="fr_%s" % (l))
            mod.update()

            # DELTA_1^l tipo 1 de bus, line l
            for l in Lines:
                delta_1[l] = mod.addVar(lb=0, vtype=GRB.BINARY, name="delta_1_%s" % (l))
            mod.update()

            # DELTA_2^l tipo 2 de bus, line l
            for l in Lines:
                delta_2[l] = mod.addVar(lb=0, vtype=GRB.BINARY, name="delta_2_%s" % (l))
            mod.update()
            #-- auxiliary variables for headway and frequency:

            # fr_k^l frequency_k for line l
            #print(list(range(1,len(frequencies_list)+1)))
            for k in range(1,len(frequencies_list)+1):
                for l in Lines:
                    fr_k[k,l]=mod.addVar(lb=0, ub=1.0, vtype=GRB.BINARY, name="fr_k_%s_%s" % (k,l))
                    #eps_k[k, l]=mod.addVar(lb=0, vtype=GRB.INTEGER, name="xi_k_%s_%s" % (k, l))
            mod.update()

            #-------- FLEET:
            # z^l required fleet
            for l in Lines:
                z[l] = mod.addVar(lb=0, vtype=GRB.INTEGER, name="z_%s" % (l))
            mod.update()

            for l in Lines:
                z_1[l] = mod.addVar(lb=0, vtype=GRB.INTEGER, name="z1_%s" % (l))
            mod.update()

            for l in Lines:
                z_2[l] = mod.addVar(lb=0, vtype=GRB.INTEGER, name="z2_%s" % (l))
            mod.update()
            #-- auxiliary variables for fleet
            for (i, j) in Lista_arcos_bus:
                 for l in Lines:
                     psi[i, j, l] = mod.addVar(lb=0, vtype=GRB.INTEGER, name="psi_%s_%s_%s" % (i, j, l))
            mod.update()

            #-- auxiliary variables for product headway and transfers:
            for (origen, destino) in dicc_parOD_listaNodos: #dicc tipo; parOD: listaNodos
                for i in dicc_parOD_listaNodos[(origen,destino)]:
                    if i != origen and i != destino:
                        for l1 in Lines:
                            for l2 in Lines:
                                if l1 != l2:
                                    eta[i,(origen, destino), l1, l2] = mod.addVar(lb=0, vtype=GRB.INTEGER, name="eta_%s_(%s,%s)_%s_%s" % (i, origen,destino, l1,l2))
            mod.update()

            #-- auxiliary variables for product headway and flow:
            for (origen, destino) in dicc_parOD_listaArcos: #dicc tipo; parOD: listaArcos
                for (i,j) in dicc_parOD_listaArcos[(origen,destino)]:
                    for l in Lines:
                        sigma[(i, j), (origen, destino), l] = mod.addVar(lb=0, vtype=GRB.INTEGER, name="rho_(%s,%s)_(%s_%s)_%s" % (i, j, origen, destino, l))
            mod.update()

            #-- auxiliary variables for product w and h:
            for (origen, destino) in dicc_parOD_listaNodos: #dicc tipo; parOD: listaNodos
                for i in dicc_parOD_listaNodos[(origen,destino)]:
                    if Lista_nodos_PED.__contains__(i):# siempre que i esté en los pedestres
                        for l in Lines:
                            rho[i, (origen, destino), l] =mod.addVar(lb=0, vtype=GRB.INTEGER, name="rho_%s_(%s,%s)_%s" % (i, origen, destino, l))


            # -----------------------------------------------------------------
            # ---------------------       CONSTRAINTS  ------------------------
            # -----------------------------------------------------------------

            # ---------   Constraint: HEADWAYS AND FREQUENCIES:
            #----------------------------------------------------------------------------------------


            # ---------   Constraint linearization of products fr_l * h^l    ------------------------

            #-------- sum_k fr_k^l=1: modificado 12/06/2023...ahora es sum_k fr_k^l=h^l:

            for l in Lines:
                mod.addConstr(quicksum(fr_k[k,l] for k in range(1,len(frequencies_list)+1)) == h[l], name="sum_fr_k_%s" % (l))
            mod.update()
            #-------- sum_k *freq_k * fr_k^l=fr^i:

            for l in Lines:
                mod.addConstr(quicksum(fr_k[k,l]* frequencies_list[k-1] for k in range(1,len(frequencies_list)+1)) == fr[l], name="sum_fr_freq_%s" % (l))
            mod.update()
            #-------- sum_k *freq_k * xi_k^l=60:   modificado 19/06/2023...ahora es sum_k *60/freq_k * fr_k^l=hw^l:

            for l in Lines:
                mod.addConstr(quicksum(fr_k[k,l] * 60/frequencies_list[k - 1] for k in range(1, len(frequencies_list) + 1)) == hw[l], name="sum_hw_freq_%s" % (l))
            mod.update()
            #
            # #-------- eps_k^l <= hw^l:
            #
            # for l in Lines:
            #     for k in range(1, len(frequencies_list) + 1):
            #         mod.addConstr(eps_k[k, l] <= hw[l], name="eps_hw_%s_%s" % (k, l))
            # mod.update()
            # #-------- eps_k^l <= 60/freq*fr_k^l:
            #
            # for l in Lines:
            #     for k in range(1, len(frequencies_list) + 1):
            #       #  mod.addConstr(eps_k[k, l] <= 60 / frequencies_list[0] * fr_k[k, l], name="eps_frk_%s_%s" % (k, l))
            #         mod.addConstr(eps_k[k, l] <= 60 / frequencies_list[k - 1] * fr_k[k, l], name="eps_frk_%s_%s" % (k, l))
            # mod.update()
            # # -------- eps_k^l >= h^l-60/freq*(1-fr_k^l):
            # for l in Lines:
            #     for k in range(1, len(frequencies_list) + 1):
            #         mod.addConstr(eps_k[k, l] >= hw[l] - 60 / frequencies_list[k - 1] * (1 - fr_k[k, l]), name="eps_frk_hw_%s_%s" % (k, l))
            #        # mod.addConstr(eps_k[k, l] >= hw[l] - 60 / frequencies_list[0] * (1 - fr_k[k, l]), name="eps_frk_hw_%s_%s" % (k, l))
            # mod.update()

            # ---------   Constraint: FLEET
            #----------------------------------------------------------------------------------------

            #-------- upperbound fleet:

            mod.addConstr(quicksum(z_1[l]  for l in Lines) <= fleet_1_max, name="sum_z1_%s<upper")
            mod.update()
            mod.addConstr(quicksum(z_2[l]  for l in Lines) <= fleet_2_max, name="sum_z2_%s<upper")
            mod.update()
            mod.addConstr(quicksum(z[l]  for l in Lines) <= fleet_max, name="sum_z_%s<upper")
            mod.update()
            for l in Lines:
                mod.addConstr(z_1[l]  <= fleet_1_max*delta_1[l], name="z1_%s<upper")
                mod.update()
                mod.addConstr(z_2[l]  <= fleet_2_max*delta_2[l], name="z2_%s<upper")
                mod.update()
                mod.addConstr(z[l] == z_1[l] +z_2[l] , name="z_equal z1_z2_")
                mod.update()
            #-------- z^l>= sum_ij (d_ij/v_bus * psi_ij^l):
                    # Otra forma con las listas;


            #Otra forma con los diccionarios;
            for l in Lines:
                mod.addConstr(z[l] >=  quicksum(
                    float(dicc_arco_distancia[(i, j)]) / v_bus * psi[i, j, l] for (i, j) in dicc_arco_distancia),
                              name="z_left_hand_%s" % (l))
                #mod.addConstr(z[l] >= quicksum(dicc_arco_distancia[(i,j)]/v_bus*psi[i,j,l]  for (i,j) in dicc_arco_distancia), name="z_left_hand_%s" % ( l))
            mod.update()

            #-------- z^l<= 1+ sum_ij (d_ij/v_bus * psi_ij^l):
            for l in Lines:
                mod.addConstr(z[l] <= 1+quicksum(float(dicc_arco_distancia[(i,j)])/v_bus*psi[i,j,l]  for (i,j) in dicc_arco_distancia), name="z_rigth_hand_%s" % ( l))
            mod.update()

            #-------- psi_ij^l <= fr^l:

            for (i,j) in Lista_arcos_bus:
                for l in Lines:
                    mod.addConstr(psi[i,j,l] <= fr[l], name="psi_fr_%s_%s_%s" % (i,j, l))
            mod.update()

            #-------- psi_ij^l <= freq_max* u_ij^l:

            for (i,j) in Lista_arcos_bus:
                for l in Lines:
                    mod.addConstr(psi[i,j,l] <= freq_max* u[i,j,l], name="psi_u_%s_%s_%s" % (i,j, l))
            mod.update()
            #-------- psi_ij^l >= fr^l-freq_max* (1-u_ij^l):

            for (i,j) in Lista_arcos_bus:
                for l in Lines:
                    mod.addConstr(psi[i,j,l] >= fr[l]-freq_max* (1-u[i,j,l]), name="psi_fr_u_%s_%s_%s" % (i,j, l))
            mod.update()

            # ---------   Constraint: Capacity constraints
            #----------------------------------------------------------------------------------------
            ##################8_05_2023---hay que ver que datos tomar, que diccionarios
            # Pasajeros viajando en arco ij de línea l
             #dicc tipo; parOD: listaArcos
            aux=0
            for l in Lines:
                for (i, j) in Lista_arcos_bus:
                    #print("*******arco", i, j)
                    for (origen, destino) in dicc_parOD_listaArcos:  # dicc tipo; parOD: listaArcos
                        #print("--parOD", origen, destino)
                        for (k, m) in dicc_parOD_listaArcos[(origen, destino)]:


                                if i == k and j==m:
                                    #print("valores:: ", (i, j), (origen, destino), l)
                                    aux=aux+r[(i, j), (origen, destino), l] * float(dicc_OD_demanda[(origen, destino)])
                                    #print("aux: ", aux)

                    mod.addConstr(aux <= fr[l] * (cap_1 * delta_1[l] + cap_2 * delta_2[l]),  name = "Restricapacidad(%s,%s)_%s" % (i, j, l))
                    #print("valor final de aux", aux)
                    aux=0
                    #print("fin de arco")

            mod.update()
            # for (origen, destino) in dicc_parOD_listaArcos: #dicc tipo; parOD: listaArcos
            #     for (i,j) in dicc_parOD_listaArcos[(origen,destino)]:
            #         for l in Lines:
            #             sigma[(i, j), (origen, destino), l] = mod.addVar(lb=0, vtype=GRB.INTEGER, name="rho_(%s,%s)_(%s_%s)_%s" % (i, j, origen, destino, l))
            # mod.update()
            # for (i,j) in Lista_arcos_bus:
            # #for (i, j) in dicc_parOD_listaArcos:
            #     for l in Lines:
            #         mod.addConstr(quicksum(r[(i, j), (origen, destino), l] *float(dicc_OD_demanda[(origen, destino)]) for (origen, destino) in dicc_parOD_listaArcos)
            #                 <= fr[l]*(cap_1*delta_1[l]+cap_2*delta_2[l]),
            #                 name="Restricapacidad(%s,%s)_%s" % (i,j, l))
            # mod.update()




            # for (origen, destino) in dicc_parOD_listaArcos: #dicc tipo; parOD: listaArcos
            #     for (i,j) in dicc_parOD_listaArcos[(origen,destino)]:
            #         for l in Lines:
            #             mod.addConstr(r[(i, j), (origen, destino), l] <=u[i,j,l], name="r_ijO^l <= u_ijl(%s,%s)_(%s,%s)_%s" % (i,j,origen, destino, l))
            #
            # mod.update()

            #-------- psi_ij^l <= delta*F_max: 06/06/2023-"@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@

            # for (i,j) in Lista_arcos_bus:
            #     for l in Lines:
            #         mod.addConstr(psi[i,j,l] <= (delta_1[l]+delta_2[l])*freq_max, name="psi_delta_%s_%s_%s" % (i,j, l))
            # mod.update()
            #-------- prueba 06/06/2023
            # for l in Lines:
            #     for (i, j) in Lista_arcos_bus:
            #         mod.addConstr(delta_1[l]+delta_2[l]>= u[i,j,l] , name="s_j_sum_u_ji%s_" % ( l))
            # mod.update()
            #--------  delta_1+delta_2<1 06/06/2023-"@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
            for l in Lines:
                mod.addConstr(delta_1[l]+delta_2[l]<=1, name="delta%s" % (l))
            mod.update()

            #Prueba tonta:
            # for l in Lines:
            #     if(l==1):
            #         for (i,j) in Lista_arcos_bus:
            #             mod.addConstr(u[i,j,l] ==0)
            #             mod.addConstr(psi[i, j, l] == 0)
            #             mod.addConstr(s[i, l]== 0)
            #             mod.addConstr(h[l] == 0)
            # mod.update()

            # for l in Lines:
            #     for (origen, destino) in dicc_parOD_listaArcos:  # dicc tipo; parOD: listaArcos
            #         for (i, j) in dicc_parOD_listaArcos[(origen, destino)]:
            #             mod.addConstr(r[(i, j), (origen, destino), l] ==0)
            #             #mod.addConstr(psi[i, j, l] == 0)
            #             #mod.addConstr(s[i, l]== 0)
            #             #mod.addConstr(h[l] == 0)
            # mod.update()

            #--------  "@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
            # ---------   Constraint: NETWORK DESIGN
            #----------------------------------------------------------------------------------------
            
            #26/2/24 RESTRICCIONES ANTIBUCLES
            #Dos nodos:
            for (i,j) in Lista_arcos_bus:
                #print("--(i,j)",i, j)
                for (k,n) in Lista_arcos_bus:                   
                    if k==j and n==i:
                        #print("--(j,i)",j, i)
                        for l in Lines:
                            mod.addConstr(u[i,j,l] + u[j,i,l] <= 1, name="antibucle 2 nodos%s_%s_%s" % (i,j, l))
            mod.update()
             #Tres nodos
            for (i,j) in Lista_arcos_bus:
               # print("**(i,j)",i, j)
                for (k,n) in Lista_arcos_bus:                   
                    if k==j:
                        for (m,o) in Lista_arcos_bus:  
                            if m==n and o==i:
                               #☻ print("**(j,i)",j, i)
                                for l in Lines:
                                    mod.addConstr(u[i,j,l] + u[k,n,l] + u[m,o,l]<= 2, name="antibucle 3 nodos%s_%s_%s" % (i,j, l))
            mod.update()
            
         #Cuatro nodos
            for (i,j) in Lista_arcos_bus:
               # print("**(i,j)",i, j)
                for (k,n) in Lista_arcos_bus:                   
                    if k==j:
                        for (m,o) in Lista_arcos_bus:  
                            if m==n:
                              for (p,q) in Lista_arcos_bus:  
                                  if p==o and q==i:
                                        for l in Lines:
                                            mod.addConstr(u[i,j,l] + u[k,n,l] + u[m,o,l]+u[o,i,l]<= 3, name="antibucle 4 nodos%s_%s_%s" % (i,j, l))
            mod.update()
            
            '''
            #Nueva restriccion 18/3/2024
            #Da problemas de infactibilidad cuando no hay flota suficiente
            
            #-------- sum_{i}{s_i^l}>=1: 
            for l in Lines:
                mod.addConstr(quicksum(s[i,l] for i in Lista_nodos)>=1, name="sum_s_i_l_>=1x%s_" % ( l))
            mod.update()
            '''
            
            #4/3/2024 RESTRICCION ENTRA=SALE
            
            for i in Lista_nodos:
                for l in Lines:
                    mod.addConstr(quicksum(u[i,k,l] for (i,k) in Lista_arcos_bus)==
                                  quicksum(u[k,i,l] for (k,i) in Lista_arcos_bus), name="sum_u_ij_l=sum_u_j_i_l%s_%s" % (j, l))
            mod.update()
            
            
            #-------- u_ij^l <= s_i^l:

            for (i,j) in Lista_arcos_bus:
                for l in Lines:
                    mod.addConstr(u[i,j,l] <= s[i,l], name="u_s_i%s_%s_%s" % (i,j, l))
            mod.update()
            #-------- u_ij^l <= s_j^l:

            for (i,j) in Lista_arcos_bus:
                for l in Lines:
                    mod.addConstr(u[i,j,l] <= s[j,l], name="u_s_j%s_%s_%s" % (i,j, l))
            mod.update()
            #-------- s_j^l =sum_{ij}{u_ij^l}:
            for j in Lista_nodos:
                for l in Lines:
                    mod.addConstr(s[j,l] ==quicksum(u[i,k,l] for (i,k) in Lista_arcos_bus if k == j), name="s_j_sum_u_ij%s_%s" % (j, l))
            mod.update()

            #-------- s_j^l =sum_{ji}{u_ji^l}:
            for j in Lista_nodos:
                for l in Lines:
                    mod.addConstr(s[j,l] ==quicksum(u[k,i,l] for (k,i) in Lista_arcos_bus if k == j), name="s_j_sum_u_ji%s_%s" % (j, l))
            mod.update()

            #-------- sum_ij{u_ij^l} =sum_{i}{s_i^l}:
            for l in Lines:
                mod.addConstr(quicksum(u[i,j,l] for (i,j) in Lista_arcos_bus) == quicksum(s[i,l] for i in Lista_nodos), name="s_j_sum_u_ji%s_" % ( l))
            mod.update()

            #-------- S_min <= sum_{i}{s_i^l}: modificada 6/6/2023#@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
            for l in Lines:
                mod.addConstr(S_min *h[l]<= quicksum(s[i,l] for i in Lista_nodos), name="S_min_sum_s_i_l%s_" % ( l))
            mod.update()
            #-------- sum_{i}{s_i^l}<= S_max: modificada 6/6/2023#@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
            for l in Lines:
                mod.addConstr(quicksum(s[i,l] for i in Lista_nodos)<=S_max*h[l], name="sum_s_i_l_S_max%s_" % ( l))
            mod.update()
            #-------- sum_{ij}{u_ij^l * dis}<= M:
            for l in Lines:
                mod.addConstr(quicksum(u[i,j,l]*dis  for (i,j,dis) in Lista_arcos_distancias_bus)<=M, name="sum_s_i_l_S_max%s_" % ( l))
            mod.update()
            #-------- r_ijO^l <= u_ijl:
            for (origen, destino) in dicc_parOD_listaArcos: #dicc tipo; parOD: listaArcos
                for (i,j) in dicc_parOD_listaArcos[(origen,destino)]:
                    for l in Lines:
                        mod.addConstr(r[(i, j), (origen, destino), l] <=u[i,j,l], name="r_ijO^l <= u_ijl(%s,%s)_(%s,%s)_%s" % (i,j,origen, destino, l))

            mod.update()

            #@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@ ---------------------6/6/2023
            for (origen, destino) in dicc_parOD_listaArcos: #dicc tipo; parOD: listaArcos
                for l1 in Lines:
                    for l2 in Lines:
                        if l1 != l2:
                            for k in dicc_parOD_listaNodos[(origen,destino)]:
                                if k != destino and k != origen:
                                    mod.addConstr(v[k, (origen, destino), l1, l2] <= 0.5*(h[l1]+h[l2]),name="nodo transfer_%s_(%s,%s)_%s_%s" % (k,origen, destino, l1,l2))
            mod.update()
            for l in Lines:
                mod.addConstr(quicksum(u[i,j,l] for (i,j) in Lista_arcos_bus)>=h[l], name="u_con h %s_" % ( l))
            mod.update()

            #@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@-----------------------fin 6/6/2023
            # ---------   Constraint: FLOW CONSERVATION
            #----------------------------------------------------------------------------------------

            # Flujo_Saliendo_Origen
            for (origen, destino) in dicc_parOD_listaArcos: #dicc tipo; parOD: listaArcos
                mod.addConstr(quicksum(quicksum(r[(i, k), (origen, destino), l] for (i, k) in dicc_parOD_listaArcos[(origen,destino)] if i == origen) for l in Lines)
                              + quicksum(t[(i,k),(origen, destino)] for (i,k) in dicc_parOD_listaAristasPED[(origen,destino)] if i == origen) == 1,
                              name="Flujo_Saliendo_Origen(%s,%s)_(%s,%s)_%s" % (i,j,origen, destino, l))
            mod.update()
            # Flujo_Entrando_Destino
            for (origen, destino) in dicc_parOD_listaArcos: #dicc tipo; parOD: listaArcos
                mod.addConstr(quicksum(quicksum(r[(i, k), (origen, destino), l] for (i, k) in dicc_parOD_listaArcos[(origen,destino)] if k == destino) for l in Lines)
                              + quicksum(t[(i,k),(origen, destino)] for (i,k) in dicc_parOD_listaAristasPED[(origen,destino)] if k == destino) == 1,
                              name="Flujo_Entrando_Destino(%s,%s)_(%s,%s)_%s" % (i,j,origen, destino, l))
            mod.update()
            #--flow balance1 .....18/06/2023
            for (origen, destino) in dicc_parOD_listaArcos: #dicc tipo; parOD: listaArcos
                for l in Lines:
                    for k in dicc_parOD_listaNodos[(origen,destino)]:#-----------------@@@@@@@@@@@@@@@@@@@@@@@@@@@la lista de aristas ped siempre debe tener una clave para cada OD
                        if k != destino and k != origen:
                            mod.addConstr(quicksum(r[(i, j), (origen, destino), l] for (i, j) in dicc_parOD_listaArcos[(origen,destino)] if j == k)
                                + quicksum(t[(i,j),(origen, destino)]  for (i,j) in dicc_parOD_listaAristasPED[(origen,destino)] if j == k)
                                          ==
                                quicksum(r[(i, j), (origen, destino), l] for (i, j) in dicc_parOD_listaArcos[(origen, destino)] if i == k)
                                          +
                                    quicksum(quicksum(r[(i, j), (origen, destino), l2] for ( i, j) in dicc_parOD_listaArcos[(origen, destino)] if i == k) for l2 in Lines if l2 != l)
                                + quicksum(t[(i, j), (origen, destino)] for (i, j) in dicc_parOD_listaAristasPED[(origen, destino)] if i == k)
                              ,name="Flujo_Balance1(%s,%s)_(%s,%s)_%s" % (i,j,origen, destino, l))
            mod.update()

            # #--flow balance2 .....18/06/2023
            # for (origen, destino) in dicc_parOD_listaArcos: #dicc tipo; parOD: listaArcos
            #     for l in Lines:
            #         for k in dicc_parOD_listaNodos[(origen,destino)]:#-----------------@@@@@@@@@@@@@@@@@@@@@@@@@@@la lista de aristas ped siempre debe tener una clave para cada OD
            #             if k != destino and k != origen:
            #                 mod.addConstr(
            #                     #quicksum(r[(i, j), (origen, destino), l] for (i, j) in dicc_parOD_listaArcos[(origen,destino)] if j == k) +
            #                     quicksum(t[(i,j),(origen, destino)]  for (i,j) in dicc_parOD_listaAristasPED[(origen,destino)] if j == k)
            #                               ==
            #                     quicksum(r[(i, j), (origen, destino), l] for (i, j) in dicc_parOD_listaArcos[(origen, destino)] if i == k)
            #                         #      +
            #                        # quicksum(quicksum(r[(i, j), (origen, destino), l2] for ( i, j) in dicc_parOD_listaArcos[(origen, destino)] if i == k) for l2 in Lines if l2 != l)
            #                     + quicksum(t[(i, j), (origen, destino)] for (i, j) in dicc_parOD_listaAristasPED[(origen, destino)] if i == k)
            #                   ,name="Flujo_Balance1(%s,%s)_(%s,%s)_%s" % (i,j,origen, destino, l))
            # mod.update()





            # ---------   Constraint: TRANSFER
            #----------------------------------------------------------------------------------------

            # transfers between lines:
            for (origen, destino) in dicc_parOD_listaArcos: #dicc tipo; parOD: listaArcos
               # pasajeros_Total_red=pasajeros_Total_red+dicc_parOD_listaArcos[(origen,destino)]
                for l1 in Lines:
                    for l2 in Lines:
                        if l1 != l2:
                            for k in dicc_parOD_listaNodos[(origen,destino)]:
                                if k != destino and k != origen:
                                    mod.addConstr(quicksum(r[(i, j), (origen, destino), l1] for (i, j) in dicc_parOD_listaArcos[(origen,destino)] if j == k)
                                       +quicksum(r[(i, j), (origen, destino), l2] for ( i, j) in dicc_parOD_listaArcos[(origen, destino)] if i == k) <=
                                                  v[k, (origen, destino), l1, l2],name="Transfer_between_lines_%s_(%s,%s)_%s_%s" % (k,origen, destino, l1,l2))
            mod.update()

            # change of modes:
            for (origen, destino) in dicc_parOD_listaArcos: #dicc tipo; parOD: listaArcos
                for l in Lines:
                    for k in dicc_parOD_listaNodos[(origen,destino)]:
                        if Lista_nodos_PED.__contains__(k):
                            mod.addConstr(quicksum(t[(i,j),(origen, destino)]  for (i,j) in dicc_parOD_listaAristasPED[(origen,destino)] if j == k)+
                                          quicksum(r[(i, j), (origen, destino), l] for (i, j) in dicc_parOD_listaArcos[(origen, destino)]
                                              if i == k)
                                          <=1 + w[k, (origen, destino), l],name="Change_of_modes%s_(%s,%s)_%s" % (k,origen, destino, l))

            mod.update()
            # v_k debe ser nodo de interseccion_1:
            for (origen, destino) in dicc_parOD_listaArcos: #dicc tipo; parOD: listaArcos
                for l1 in Lines:
                    for l2 in Lines:
                        if l1 != l2:
                            for k in dicc_parOD_listaNodos[(origen,destino)]:
                                if k != origen and k != destino:
                                    mod.addConstr( v[k, (origen, destino), l1, l2] <=s[k,l1],name="v_k_nodoInterseccion_1%s_(%s,%s)_%s_%s" % (k,origen, destino, l1,l2))

            mod.update()
            # v_k debe ser nodo de interseccion_2:
            for (origen, destino) in dicc_parOD_listaArcos: #dicc tipo; parOD: listaArcos
                for l1 in Lines:
                    for l2 in Lines:
                        if l1 != l2:
                            for k in dicc_parOD_listaNodos[(origen,destino)]:
                                if k != origen and k != destino:
                                    mod.addConstr( v[k, (origen, destino), l1, l2] <=s[k,l2],name="v_k_nodoInterseccion_2%s_(%s,%s)_%s_%s" % (k,origen, destino, l1,l2))

            mod.update()



            # -----------------------------------------------------------------
            # ------------      OBJECTIVE FUNCTION         --------------------
            # -----------------------------------------------------------------
            z_bus = mod.addVar(lb=0, ub=GRB.INFINITY, vtype=GRB.CONTINUOUS, name="z_bus")
            z_walk = mod.addVar(lb=0, ub=GRB.INFINITY, vtype=GRB.CONTINUOUS, name="z_walk")
            z_trans = mod.addVar(lb=0, ub=GRB.INFINITY, vtype=GRB.CONTINUOUS, name="z_trans")
            z_wait = mod.addVar(lb=0, ub=GRB.INFINITY, vtype=GRB.CONTINUOUS, name="z_wait")
            z_change = mod.addVar(lb=0, ub=GRB.INFINITY, vtype=GRB.CONTINUOUS, name="z_change")
            f_objetivo=mod.addVar(lb=0, ub=GRB.INFINITY, vtype=GRB.CONTINUOUS, name="f_obj")
            #----- z_bus=60/v_bus*sum_OD{sum_l{sum_ij{d_ij*r_ij^Ol*p_O
            # demanda_theta=float(dicc_OD_demanda[(origen, destino)])
            # distancia=float(dicc_arco_distancia[(i, j)]
            mod.addConstr(z_bus == 60/v_bus*
                          quicksum(quicksum(quicksum(r[(i, j), (origen, destino), l]*float(dicc_OD_demanda[(origen, destino)])*float(dicc_arco_distancia[(i, j)])
                                                     for (i,j) in dicc_parOD_listaArcos[(origen, destino)] )
                                                            for l in Lines) for (origen, destino) in dicc_parOD_listaArcos), name="z_bus")

            mod.update()

            #----- z_walk=60/v_PED*sum_OD{sum_ij{d_ij^PED*t_ij^O*p_O
            # demanda_theta=float(dicc_OD_demanda[(origen, destino)])
            # distancia=float(dicc_arco_distancia[(i, j)]
            mod.addConstr(z_walk == 60/v_PED*
                          quicksum(quicksum(t[(i, j), (origen, destino)]*float(dicc_OD_demanda[(origen, destino)])*float(dicc_arco_distancia_PED[(i, j)])
                                                     for (i,j) in dicc_parOD_listaAristasPED[(origen, destino)] )
                                                             for (origen, destino) in dicc_parOD_listaAristasPED), name="z_walk")

            mod.update()



            #----- z_trans=sum_OD{sum_i{sum_l{eta_i^Oll'*h^l/2*p_O
            # demanda_theta=float(dicc_OD_demanda[(origen, destino)])
            # distancia=float(dicc_arco_distancia[(i, j)]
            mod.addConstr(z_trans == quicksum(quicksum(quicksum(quicksum(
                eta[i, (origen, destino), l1,l2]*float(dicc_OD_demanda[(origen, destino)])/2 for l2 in Lines if l2 !=l1)
                                                                for l1 in Lines)
                                                                    for i in dicc_parOD_listaNodos[(origen, destino)]  if i != origen and i != destino)
                                                                            for (origen, destino) in dicc_parOD_listaNodos), name="z_trans")

            mod.update()


            #----- z_wait=sum_OD{sum_(O,j){sum_l{sigma^0l*p_O/2
            mod.addConstr(z_wait ==
                          quicksum(quicksum(quicksum(sigma[(i, j), (origen, destino), l] * float(dicc_OD_demanda[(origen, destino)]) / 2
                                                     for l in Lines)
                                                     for (i,j) in dicc_parOD_listaArcos[(origen, destino)] if i ==origen)
                                                            for (origen, destino) in dicc_parOD_listaArcos), name="z_wait")

            mod.update()



            #----- z_change=sum_OD{sum_(O,j){sum_l{rho_i^0l*p_O/2
            mod.addConstr(z_change ==quicksum(quicksum(quicksum(rho[i, (origen, destino), l] * float(dicc_OD_demanda[(origen, destino)]) / 2
                                                     for l in Lines)
                                                        for i in dicc_parOD_listaNodos[(origen, destino)] if Lista_nodos_PED.__contains__(i))
                                                            for (origen, destino) in dicc_parOD_listaNodos), name="z_change")

            mod.update()



            # --------------------------
            # ---------   Constraint linearization of products v_i Ôll' * h^l    ------------------------

            #-------- eta_i^Oll' <= hw^l':

            for l1 in Lines:
                for l2 in Lines:
                    if l1 != l2:
                        for (origen, destino) in dicc_parOD_listaNodos:
                            for i in dicc_parOD_listaNodos[(origen, destino)]:
                                if i!= origen and i != destino:
                                    mod.addConstr(eta[i,(origen,destino), l1, l2] <= hw[l2], name="eta_h_%s_(%s,%s)_%s_%s" % (i, origen, destino, l1,l2))
            mod.update()

            #-------- eta_i^Oll' <= 60/freq_max*v_i ^Oll':

            for l1 in Lines:
                for l2 in Lines:
                    if l1 != l2:
                        for (origen, destino) in dicc_parOD_listaNodos:
                            for i in dicc_parOD_listaNodos[(origen, destino)]:
                                if i!= origen and i != destino:
                                    mod.addConstr(eta[i,(origen,destino), l1, l2] <= 60/freq_min*v[i, (origen, destino), l1, l2], name="eta_h_%s_(%s,%s)_%s_%s" % (i, origen, destino, l1,l2))
                             #13/03/2023       mod.addConstr(eta[i,(origen,destino), l1, l2] <= 60/freq_max*v[i, (origen, destino), l1, l2], name="eta_h_%s_(%s,%s)_%s_%s" % (i, origen, destino, l1,l2))
            mod.update()
            #-------- eta_i^Oll' >= h^l'- 60/freq_max*(1-v_i ^Oll'):

            for l1 in Lines:
                for l2 in Lines:
                    if l1 != l2:
                        for (origen, destino) in dicc_parOD_listaNodos:
                            for i in dicc_parOD_listaNodos[(origen, destino)]:
                                if i!= origen and i != destino:
                                   #13/03/2023  mod.addConstr(eta[i,(origen,destino), l1, l2] >= hw[l2]*60/freq_max*(1-v[i, (origen, destino), l1, l2]), name="eta_h_%s_(%s,%s)_%s_%s" % (i, origen, destino, l1,l2))
                                    mod.addConstr(eta[i, (origen, destino), l1, l2] >= hw[l2] - 60 / freq_min * (
                                                1 - v[i, (origen, destino), l1, l2]),
                                                  name="eta_h_%s_(%s,%s)_%s_%s" % (i, origen, destino, l1, l2))
            mod.update()

            # ---------   Constraint linearization of products r_ij Ôl * h^l    ------------------------
            #-------- sigma_ij^Ol <= hw^l:

            for (origen, destino) in dicc_parOD_listaArcos: #dicc tipo; parOD: listaArcos
                for (i,j) in dicc_parOD_listaArcos[(origen,destino)]:
                        for l in Lines:
                                    mod.addConstr(sigma[(i, j), (origen, destino), l] <= hw[l], name="sigma_h_(%s,%s)_(%s,%s)_%s" % (i, j, origen, destino, l))
            mod.update()
            #-------- sigma_ij^Ol <= 60/freq_max*r_ij^Ol:

            for (origen, destino) in dicc_parOD_listaArcos: #dicc tipo; parOD: listaArcos
                for (i,j) in dicc_parOD_listaArcos[(origen,destino)]:
                        for l in Lines:
                                    # 13/03/2023 mod.addConstr(sigma[(i, j), (origen, destino), l] <= 60 / freq_max * r[(i, j), (origen, destino), l], name="sigma_r_(%s,%s)_(%s,%s)_%s" % (i, j, origen, destino, l))
                                    mod.addConstr(sigma[(i, j), (origen, destino), l] <= 60 / freq_min * r[(i, j), (origen, destino), l],
                                      name="sigma_r_(%s,%s)_(%s,%s)_%s" % (i, j, origen, destino, l))
            mod.update()
            #-------- sigma_ij^Ol >=h^l- 60/freq_max*(1-r_ij^Ol):

            for (origen, destino) in dicc_parOD_listaArcos: #dicc tipo; parOD: listaArcos
                for (i,j) in dicc_parOD_listaArcos[(origen,destino)]:
                        for l in Lines:
                                    mod.addConstr(sigma[(i, j), (origen, destino), l] >= hw[l] - 60 / freq_min * (1 - r[(i, j), (origen, destino), l]), name="sigma_h_r_(%s,%s)_(%s,%s)_%s" % (i, j, origen, destino, l))
                                    # 13/03/2023  mod.addConstr(sigma[(i, j), (origen, destino), l] >= hw[l] - 60 / freq_max * (1 - r[(i, j), (origen, destino), l]), name="sigma_h_r_(%s,%s)_(%s,%s)_%s" % (i, j, origen, destino, l))
            mod.update()

            # ---------   Constraint linearization of products w_i^Ol * h^l    ------------------------
            #----- rho_i^Ol <= w_i^Ol

            for (origen, destino) in dicc_parOD_listaNodos:
                for i in dicc_parOD_listaNodos[(origen, destino)]:
                    if Lista_nodos_PED.__contains__(i):
                        mod.addConstr(rho[i, (origen, destino), l] <= hw[l], name="rho_h_%s(%s,%s)_%s" % (i, origen, destino, l))
            mod.update()
            #----- rho_i^Ol <= 60/freq_max* w_i^Ol

            for (origen, destino) in dicc_parOD_listaNodos:
                for i in dicc_parOD_listaNodos[(origen, destino)]:
                    if Lista_nodos_PED.__contains__(i):
                        mod.addConstr(rho[i, (origen, destino), l] <= 60/freq_min*w[i, (origen, destino), l], name="rho_w_%s(%s,%s)_%s" % (i, origen, destino, l))
                        # 13/03/2023 mod.addConstr(rho[i, (origen, destino), l] <= 60 / freq_max * w[i, (origen, destino), l], name="rho_w_%s(%s,%s)_%s" % (i, origen, destino, l))
            mod.update()
            #----- rho_i^Ol >= h^l-60/freq_max* (1-w_i^Ol)

            for (origen, destino) in dicc_parOD_listaNodos:
                for i in dicc_parOD_listaNodos[(origen, destino)]:
                    if Lista_nodos_PED.__contains__(i):
                        # 13/03/2023 mod.addConstr(rho[i, (origen, destino), l] >= hw[l]-60/freq_max*(1-w[i, (origen, destino), l]), name="rho_h_w_%s(%s,%s)_%s" % (i, origen, destino, l))
                        mod.addConstr(rho[i, (origen, destino), l] >= hw[l]-60/freq_min*(1-w[i, (origen, destino), l]), name="rho_h_w_%s(%s,%s)_%s" % (i, origen, destino, l))
            mod.update()

            # ---------  END CREATION OF OBJECTIVE FUNCTION    ------------------------

            # ---------------------------------------------------------------------------------
            # ---------   LLAMADA AL SOLVER   ------------------------
            # ---------------------------------------------------------------------------------
            mod.addConstr(f_objetivo == z_trans+z_bus+z_change+z_wait+z_walk, name="travel_time")
            mod.setObjective(f_objetivo, sense=GRB.MINIMIZE)
            mod.Params.MIPGap = 0.010
            #mod.Params.TimeLimit = 360
            #mod.setParam('TIME_LIMIT', 5)

            #mod.setParam('MIPGap', 0.05)
            #mod.setParam('Runtime', 30)
            #mod.setParam('Timelimit', 30)

            mod.write('Buses.lp')#imprime el modelo tal cual
            #mylogfile=r'C:\Users\David1\OneDrive - UNIVERSIDAD DE SEVILLA\PyCharmProjects\BUS\mylog.log'
            #mod.params.LogFile = mylogfile
            print("#####################################################-----------------------------RESOLVIENDO ESCENARIO: ", contadorEscenario)
            mod.optimize()

            #---------------------------------------------------------------------------------
            # ---------   RECOGIDA DE LA INFORMACIÓN DE LA SOLUCIÓN   ------------------------
            #---------------------------------------------------------------------------------
            print("Trazassss")
            num_pares_red=0
            num_viajes_directos=0
            num_transbordos=0
            num_pares_con_transfer=0
            tiempo_bus = 0
            tiempo_walk = 0
            tiempo_transfer = 0
            tiempo_wait = 0
            tiempo_change = 0
            contadorColumnas_escenario = 0
            pasajeros_Total_red=0
            pasajerosLinea=0

            # Cálculo del numpasajeros_Total_red y num_pares_red:
            for (origen, destino) in dicc_OD_demanda:  # dicc tipo; parOD: listaArcos
                pasajeros_Total_red = pasajeros_Total_red + float(dicc_OD_demanda[(origen, destino)])
                num_pares_red=num_pares_red+1
                # for l in Lines:
                #     for l2 in Lines:
                #         if h[l].x > 0:

            #Cálculo del num transbordos total:
            state=True
            for l1 in Lines:

                for l2 in Lines:
                    if l1 != l2:
                       # if l1 != l2 & h[l1].x > 0:
                        for (origen, destino) in dicc_parOD_listaNodos:
                           # while state==True:
                            for i in dicc_parOD_listaNodos[(origen, destino)]:
                                if i != origen and i != destino:
                                    if v[i, (origen, destino), l1,l2].x>0 and int(dicc_OD_demanda[(origen, destino)]) > 0:
                                        num_transbordos=num_transbordos+ v[i, (origen, destino), l1,l2].x
                                           # state=False

            # Cálculo del num pares con transf¡bordos :
            state = True
            for l1 in Lines:
              
                       
                for l2 in Lines:
                    if l1 != l2  and h[l1].x > 0 and h[l2].x > 0:
                        # if l1 != l2 & h[l1].x > 0:
                        for (origen, destino) in dicc_parOD_listaNodos:
                            state = True
                            while state:
                                for i in dicc_parOD_listaNodos[(origen, destino)]:
                                    if i != origen and i != destino:
                                       
                                        if v[i, (origen, destino), l1, l2].x > 0 and int(dicc_OD_demanda[(origen, destino)]) > 0:
                                            num_pares_con_transfer = num_pares_con_transfer + v[i, (origen, destino), l1, l2].x
                                            print("Transbordo: ", "parada: ",i, "par: ", (origen, destino), "lineas: ", (l1, l2))
                                            state=False
                                state=False
                                
        # Cálculo de pares andando
            num_pares_andando=0
            xx=0
            yy=0
    
            for (origen, destino) in dicc_parOD_listaArcos:
                xx=0
                yy=0
                if int(dicc_OD_demanda[(origen, destino)]) > 0:
                    for l in Lines:                  
                            
                        for (i, j) in dicc_parOD_listaArcos[(origen, destino)]:
                               
                            xx=xx+r[(i, j), (origen, destino), l].x   
                    if xx==0  :
                        num_pares_andando=num_pares_andando+1
                   
            # Cálculo de pares bus 
            num_pares_bus=0
            xx=0
            yy=0
            numero_pares_OD=0
            for (origen, destino) in dicc_parOD_listaArcos:
                if int(dicc_OD_demanda[(origen, destino)]) > 0:
                    numero_pares_OD=numero_pares_OD+1
                    xx=0
                    yy=0
                    for l in Lines:                  
                            
                        for (i, j) in dicc_parOD_listaArcos[(origen, destino)]:
                               
                            xx=xx+r[(i, j), (origen, destino), l].x   
                    if xx>0  :
                        num_pares_bus=num_pares_bus+1
                        
                                
                       #     state=False                                   
            print("numero pares",numero_pares_OD)
            num_viajes_directos=numero_pares_OD-num_pares_con_transfer-num_pares_andando
            print("num_pares_red", num_pares_red)
            print("num_pares_andando", num_pares_andando)
            print("num_pares_con_transfer", num_pares_con_transfer, num_transbordos)
            print("num_pares_bus", num_pares_bus)
            # Cálculo de z_bus:
            for l in Lines:
                for (origen, destino) in dicc_parOD_listaArcos:
                    for (i, j) in dicc_parOD_listaArcos[(origen, destino)]:
                        if r[(i, j), (origen, destino), l].x > 0 and int(dicc_OD_demanda[(origen, destino)]) > 0:
                            print("Arco: ", (i, j), "Par: ", (origen, destino), "linea: ",l)
                        tiempo_bus = tiempo_bus + 60 / v_bus * r[(i, j), (origen, destino), l].x * float(
                            dicc_OD_demanda[(origen, destino)]) * float(dicc_arco_distancia[(i, j)])
            # Cálculo de z_walk
            for l in Lines:
                for (origen, destino) in dicc_parOD_listaAristasPED:
                    for (i, j) in dicc_parOD_listaAristasPED[(origen, destino)]:
                        tiempo_walk = tiempo_walk + 60 / v_PED * t[(i, j), (origen, destino)].x * float(
                            dicc_OD_demanda[(origen, destino)]) * float(dicc_arco_distancia_PED[(i, j)])
                        
                        

            # Cálculo de z_transfer
            for l1 in Lines:
                for l2 in Lines:
                    if l1 != l2:
                        for (origen, destino) in dicc_parOD_listaNodos:
                            for i in dicc_parOD_listaNodos[(origen, destino)]:
                                if i != origen and i != destino:
                                    tiempo_transfer = tiempo_transfer + eta[i, (origen, destino), l1, l2].x * float(
                                        dicc_OD_demanda[(origen, destino)]) / 2
                                    # if v[i, (origen, destino), l1, l2].x>0:
                                    #         if hw[l2].x>0:
                                    #             if eta[i, (origen, destino), l1, l2].x>0:
                                    #                 print("TODO OKKK")
                                    #             else:
                                    #                 print("HAY ALGO RARO")

            # Cálculo de z_wait
            for l in Lines:
                for (origen, destino) in dicc_parOD_listaArcos:
                    for (i, j) in dicc_parOD_listaArcos[(origen, destino)]:
                        if i == origen:
                            tiempo_wait = tiempo_wait + sigma[(i, j), (origen, destino), l].x * float(dicc_OD_demanda[(origen, destino)]) / 2


            # Cálculo de change
            for l1 in Lines:
                for (origen, destino) in dicc_parOD_listaNodos:
                    for i in dicc_parOD_listaNodos[(origen, destino)]:
                        if Lista_nodos_PED.__contains__(i):
                            tiempo_change=tiempo_change+ rho[i, (origen, destino), l].x * float(dicc_OD_demanda[(origen, destino)]) / 2

            #---------------------------------------------------------------------------------
            ######          CREACION FECHERO EXCEL SOLUCION   --------------------------------
            #---------------------------------------------------------------------------------

            #Creamos dos ficheros; uno con la información del escenario y otro con las soluciones globales

            #Creación fichero excel con la solucion global de la red
            workbook_Solution_Global=xlwt.Workbook()
            workbook_Sheet_Global=workbook_Solution_Global.add_sheet(instance) #instance = nombre de la pestaña

            #Creación fichero excel con la solucion del escenario
            # workbook_Solution_Escenario=openpyxl.Workbook()
            # workbook_Sheet_Escenario=workbook_Solution_Escenario.create_sheet(instance) #instance = nombre de la pestaña

            workbook_Solution_Escenario=xlwt.Workbook()
            workbook_Sheet_Escenario=workbook_Solution_Escenario.add_sheet(instance) #instance = nombre de la pestaña
            workbook_Sheet_Escenario_descripcion = workbook_Solution_Escenario.add_sheet(str(contadorEscenario))  # instance = nombre de la pestaña
            #------------cabeceras de la tabla----------------------
            workbook_Sheet_Escenario.write(0,0,"Línea activada") #workbook_Sheet_Escenario.write(fila,col,"Línea activada")
            workbook_Sheet_Escenario.write(0,1,"Nb nodos")
            workbook_Sheet_Escenario.write(0,2,"Nb arcos")
            workbook_Sheet_Escenario.write(0,3,"Lista nodos")
            workbook_Sheet_Escenario.write(0,4,"Lista arcos")
            workbook_Sheet_Escenario.write(0,5,"Nb autobuses")
            workbook_Sheet_Escenario.write(0,6,"Tipo autobus")
            workbook_Sheet_Escenario.write(0,7,"Frecuencia")
            workbook_Sheet_Escenario.write(0,8,"Longitud")
            workbook_Sheet_Escenario.write(0,9,"Pasajeros que usan línea")
            workbook_Sheet_Escenario.write(0,10,"Nb pares que usan línea")


            #--------------Recogida de datos de la solución............
            contadorFilas=1
            contadorColumnas_escenario=0

            #Iintroducimos las cabeceras de los datos del escenario:
            workbook_Sheet_Escenario_descripcion.write(0,0,"Num max de líneas") #
            workbook_Sheet_Escenario_descripcion.write(0, 1, "Lista de frecuencias")  #
            workbook_Sheet_Escenario_descripcion.write(0, 2, "Num max flota")  #
            workbook_Sheet_Escenario_descripcion.write(0, 3, "Num max flota Tipo 1")  #
            workbook_Sheet_Escenario_descripcion.write(0, 4, "Num max flota Tipo 2")  #
            workbook_Sheet_Escenario_descripcion.write(0, 5, "Pasajeros de la red")  #
            workbook_Sheet_Escenario_descripcion.write(0, 6, "Num pares de la red")  #
            workbook_Sheet_Escenario_descripcion.write(0, 7, "GAP")  #
            workbook_Sheet_Escenario_descripcion.write(0, 8, "CPU")  #
            workbook_Sheet_Escenario_descripcion.write(0, 9, "Num viajes directos")  #
            workbook_Sheet_Escenario_descripcion.write(0, 10, "Num transbordos")  #
            workbook_Sheet_Escenario_descripcion.write(0, 11, "tiempo en bus (z_bus)")  #
            workbook_Sheet_Escenario_descripcion.write(0, 12, "tiempo andando (z_walk)")  #
            workbook_Sheet_Escenario_descripcion.write(0, 13, "tiempo transbordando (z_transfer)")  #
            workbook_Sheet_Escenario_descripcion.write(0, 14, "tiempo esperando (z_wait)")  #
            workbook_Sheet_Escenario_descripcion.write(0, 15, "tiempo espera cambio modos (z_change)")  #
            workbook_Sheet_Escenario_descripcion.write(0, 16, "valor de la f.o.")
            #...................Introducimos los datos del escenario

            workbook_Sheet_Escenario_descripcion.write(1,contadorColumnas_escenario,numLineas) #
            contadorColumnas_escenario=contadorColumnas_escenario+1
            workbook_Sheet_Escenario_descripcion.write(1,contadorColumnas_escenario,str(frequencies_list)) #
            contadorColumnas_escenario=contadorColumnas_escenario+1
            workbook_Sheet_Escenario_descripcion.write(1,contadorColumnas_escenario,fleet_max) #
            contadorColumnas_escenario=contadorColumnas_escenario+1
            workbook_Sheet_Escenario_descripcion.write(1,contadorColumnas_escenario,fleet_1_max) #
            contadorColumnas_escenario=contadorColumnas_escenario+1
            workbook_Sheet_Escenario_descripcion.write(1,contadorColumnas_escenario,fleet_2_max) #
            contadorColumnas_escenario=contadorColumnas_escenario+1
            workbook_Sheet_Escenario_descripcion.write(1, contadorColumnas_escenario, pasajeros_Total_red)  #
            contadorColumnas_escenario = contadorColumnas_escenario + 1
            workbook_Sheet_Escenario_descripcion.write(1, contadorColumnas_escenario, num_pares_red)  #
            contadorColumnas_escenario = contadorColumnas_escenario + 1

            workbook_Sheet_Escenario_descripcion.write(1, contadorColumnas_escenario, mod.MIPGap)  #
            contadorColumnas_escenario = contadorColumnas_escenario + 1
            workbook_Sheet_Escenario_descripcion.write(1, contadorColumnas_escenario, mod.Runtime)  #
            contadorColumnas_escenario = contadorColumnas_escenario + 1

            workbook_Sheet_Escenario_descripcion.write(1, contadorColumnas_escenario, num_viajes_directos)  #
            contadorColumnas_escenario = contadorColumnas_escenario + 1
            workbook_Sheet_Escenario_descripcion.write(1, contadorColumnas_escenario, num_transbordos)  #
            contadorColumnas_escenario = contadorColumnas_escenario + 1


            # RECOGEMOS LOS DATOS DE LA SOLUCIÓN PARA INCLUIRLAS EN EL EXCEL:
            print("valor de la f.o.", str(mod.ObjVal))
            listaNodosLinea={}

            Solarco = []
            Solnodo = []
            nodoList=[]
            arcoList=[]
            FlowLinksSol = {}
            TransferNodesol = {}
            FlowPerSol = {}
            Flota_total=0
            
            #5/3/2024 modificaciones para dibujar grafo
            Lista_Lineas=[]


            for l in Lines:
                contadorColumnas=0
                pasajerosLinea = 0
                NbParesLinea=0
                longitud_linea=0
                suma_flota_min_psi=0
                suma_flota_min = 0
                suma_flota_max = 0
                nodoList = []
                arcoList = []
                print("**************************************************")
                print("Line %s:" % (l))
                print("**************************************************")
                if h[l].x > 0:
                    print("++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++linea se activa %s" % (l))

                    workbook_Sheet_Escenario.write(contadorFilas, contadorColumnas, l)
                    contadorColumnas=contadorColumnas+1
                    for i in Lista_nodos:
                        if s[i, l].x > 0:
                            Solnodo.append((i, l))
                            nodoList.append(i)
                            print("Nodo %s" % (i))
                    print("--------------------------------------------------")
                    #print("nodoList: ", nodoList)
                    #print(str(len(nodoList)))

                    workbook_Sheet_Escenario.write(contadorFilas, contadorColumnas, str(len(nodoList)))# Nodos

                    contadorColumnas=contadorColumnas+1
                    for (i, j) in Lista_arcos_bus:
                        if u[i, j, l].x > 0:
                            Solarco.append((i, j, l))
                            arcoList.append((i,j))
                            #print("Arco (%s,%s)" % (i, j))
                    workbook_Sheet_Escenario.write(contadorFilas, contadorColumnas, len(arcoList))#Arcos
                    contadorColumnas=contadorColumnas+1
                    Lista_Lineas.append(arcoList)
                    workbook_Sheet_Escenario.write(contadorFilas, contadorColumnas, str(nodoList))  # lista de nodos
                    contadorColumnas=contadorColumnas+1

                    workbook_Sheet_Escenario.write(contadorFilas, contadorColumnas, str(arcoList))  # lista de arcos
                    contadorColumnas=contadorColumnas+1
                    if z[l].x > 0:
                        #print("Flota z%s" % (l), "numero buses", z[l].x)
                        Flota_total=Flota_total+ z[l].x
                        workbook_Sheet_Escenario.write(contadorFilas, contadorColumnas, z[l].x)  # Nb autobuses
                        contadorColumnas=contadorColumnas+1
                    if delta_1[l].x == 1:
                        #print("Delta 1 %s" % (l), "bus tipo 1")
                        #print("Flota z1 %s" % (l), "bus tipo 1", z_1[l].x)
                        workbook_Sheet_Escenario.write(contadorFilas, contadorColumnas, 'Tipo 1')  # Tipo autobus 1
                        contadorColumnas=contadorColumnas+1
                    if delta_2[l].x == 1:
                        #print("Delta 2 %s" % (l), "bus tipo 2")
                        #print("Flota z2 %s" % (l), "bus tipo 1", z_2[l].x)
                        workbook_Sheet_Escenario.write(contadorFilas, contadorColumnas, 'Tipo 2')  # Tipo autobus 2
                        contadorColumnas=contadorColumnas+1
                    if fr[l].x > 0:
                        #print("Frecuencia linea:", fr[l].x)
                        workbook_Sheet_Escenario.write(contadorFilas, contadorColumnas, fr[l].x)  # Frecuencia
                        contadorColumnas=contadorColumnas+1
                        #print("headway linea:", hw[l].x)

                    for (i, j) in dicc_arco_distancia:
                        if u[i, j, l].x > 0:
                            suma_flota_min=suma_flota_min+float(dicc_arco_distancia[(i, j)]) / v_bus * u[i, j, l].x *fr[l].x
                            suma_flota_min_psi = suma_flota_min_psi + float(dicc_arco_distancia[(i, j)]) / v_bus * psi[i, j, l].x
                            longitud_linea = longitud_linea + float(dicc_arco_distancia[(i, j)]) * u[i, j, l].x
                            #print("Flota total en la red: ", Flota_total)

                            #print("longitud_linea: ", longitud_linea)
                    workbook_Sheet_Escenario.write(contadorFilas, contadorColumnas, longitud_linea)  # Longitud de la linea

                    contadorColumnas=contadorColumnas+1

                    # Cálculo pasajeros de la línea y número de pares atendidos por la línea

                    for (origen, destino) in dicc_OD_demanda:
                        for (i, j) in dicc_parOD_listaArcos[(origen, destino)]:
                            if r[(i, j), (origen, destino), l].x > 0:
                                float(dicc_OD_demanda[(origen, destino)])
                                pasajerosLinea = pasajerosLinea + float(dicc_OD_demanda[(origen, destino)])
                                NbParesLinea = NbParesLinea + 1
                                break
                   #print("dicc_parOD_listaNodos:", dicc_parOD_listaNodos)
                    # for (origen, destino) in dicc_OD_demanda:
                    #     for i in dicc_parOD_listaNodos[(origen, destino)]:
                    #         for l2 in Lines:
                    #             if l != l2:
                    #                 if i != origen and i != destino:
                                       # print("ETA", eta[i, (origen, destino), l, l2].x)


                    workbook_Sheet_Escenario.write(contadorFilas, contadorColumnas, pasajerosLinea)  # Pasajeros

                    contadorColumnas=contadorColumnas+1
                    workbook_Sheet_Escenario.write(contadorFilas, contadorColumnas, NbParesLinea)  # Nb pares
                    contadorColumnas=contadorColumnas+1

                contadorFilas=contadorFilas+1


            print("--------------------------------------------------")
            #................
            workbook_Sheet_Escenario_descripcion.write(1, contadorColumnas_escenario, tiempo_bus)  #
            contadorColumnas_escenario = contadorColumnas_escenario + 1
            #workbook_Sheet_Escenario_descripcion.write(1, contadorColumnas_escenario, tiempo_walk)  #z_walk
            workbook_Sheet_Escenario_descripcion.write(1, contadorColumnas_escenario, z_walk.x)  #
            contadorColumnas_escenario = contadorColumnas_escenario + 1
            #workbook_Sheet_Escenario_descripcion.write(1, contadorColumnas_escenario, tiempo_transfer)  #
            workbook_Sheet_Escenario_descripcion.write(1, contadorColumnas_escenario, z_trans.x)  #



            contadorColumnas_escenario = contadorColumnas_escenario + 1
#            workbook_Sheet_Escenario_descripcion.write(1, contadorColumnas_escenario, tiempo_wait)  #
            workbook_Sheet_Escenario_descripcion.write(1, contadorColumnas_escenario, z_wait.x)  #
            contadorColumnas_escenario = contadorColumnas_escenario + 1
            workbook_Sheet_Escenario_descripcion.write(1, contadorColumnas_escenario, z_change.x)  #
            #workbook_Sheet_Escenario_descripcion.write(1, contadorColumnas_escenario, tiempo_change)  #
            contadorColumnas_escenario = contadorColumnas_escenario + 1
            workbook_Sheet_Escenario_descripcion.write(1, contadorColumnas_escenario, mod.ObjVal)  #
            contadorColumnas_escenario = contadorColumnas_escenario + 1
            
            #5/3/2024 Añadir grafo al excel ---------------------------
            
            grafo = GrafoColores(Lista_Lineas)
            grafo.dibujar_grafo()
            image_path = grafo.image_path
            
            
            
                                    
            # SE CIERRA EL EXCEL CON LA SOLUCIÓN DEL ESCENARIO

            workbook_Solution_Escenario.save("C:/Users/Fernando/Desktop/TFG/pythonMathModels/Soluciones/"+instance+'_Escenario_'+str(contadorEscenario)+'.xls')
            #workbook_Solution_Escenario.save(pathSoluciones+instance+'_Escenario_'+contadorEscenario+'.xls')
            #"C:/Users/aliciasantos/Documents/EXPERIMENTOS/Proyectos Python/pythonMathModels/Soluciones/",

            #workbook_Solution.save('C:/proyectos Python/pythonMeasures/Soluciones/Solucion_'+instance+'.xls')

            # Paso 1: Lee todas las hojas del archivo .xls con pandas
            dfs = pd.read_excel("C:/Users/Fernando/Desktop/TFG/pythonMathModels/Soluciones/"+instance+'_Escenario_'+str(contadorEscenario)+'.xls', sheet_name=None)
            
            # Paso 2: Escribe cada DataFrame en un archivo .xlsx
            with pd.ExcelWriter("C:/Users/Fernando/Desktop/TFG/pythonMathModels/Soluciones/SolucionesDibujo/"+instance+'_Escenario_'+str(contadorEscenario)+'.xlsx') as writer:
                for sheet_name, df in dfs.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                        # Carga el libro de trabajo existente
            book = load_workbook("C:/Users/Fernando/Desktop/TFG/pythonMathModels/Soluciones/SolucionesDibujo/"+instance+'_Escenario_'+str(contadorEscenario)+'.xlsx')
            
            # Selecciona la hoja donde quieres añadir la imagen
            sheet = book['Mandl']
            
            # Carga la imagen
            img = Image(image_path)
            
            # Añade la imagen a la hoja en la celda 'A20'
            sheet.add_image(img, 'N1')
            
            # Guarda el libro de trabajo
            book.save("C:/Users/Fernando/Desktop/TFG/pythonMathModels/Soluciones/SolucionesDibujo/"+instance+'_Escenario_'+str(contadorEscenario)+'.xlsx')


            #................................................................

            #mod.write('inisolution.mst')

            # for v in mod.getVars():
            #     if v.x!=0:
            #
            #         #print("Variable " + str(v.VarName) + " = " + str(v.x))
            #         var_str=str(v.VarName)
            #         variable = var_str.split("_")
            #         name = variable[0]
            #         resto = variable[1]
            #        # print("valor,", variable)
            #         #print("name,", name)
            #         #print("resto", resto)
            #         if name=="s":
            #             print('%s %s' % (v.VarName,v.x))
            #             # i=resto[0]
            #             # l=resto[0]
            #             # listaNodosLinea[0].a
            #
            # print(listaNodosLinea)
            #mod.write('inisolution.mst')
            status = mod.Status
            print('status',status)
            #print (model.display())